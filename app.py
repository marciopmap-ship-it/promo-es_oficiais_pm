from __future__ import annotations

from dataclasses import dataclass, asdict
from io import BytesIO
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REQUIRED_COLUMNS = ["Nome", "Antiguidade", "Merecimento", "Agregado"]


@dataclass
class Promotion:
    ordem: int
    nome: str
    criterio: str
    agregado: str
    consumiu_vaga: str
    vaga_principal: int
    observacao: str


def normalize_bool(value: Any) -> bool:
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    return text in {
        "sim",
        "s",
        "true",
        "t",
        "1",
        "yes",
        "y",
        "agregado",
    }


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Colunas obrigatorias ausentes: {', '.join(missing)}")

    df = df[REQUIRED_COLUMNS].copy()
    df["Nome"] = df["Nome"].astype(str).str.strip()
    df = df[df["Nome"] != ""]
    df["Antiguidade"] = pd.to_numeric(df["Antiguidade"], errors="raise")
    df["Merecimento"] = pd.to_numeric(df["Merecimento"], errors="raise")
    df["Agregado"] = df["Agregado"].apply(normalize_bool)

    duplicated = df[df["Nome"].duplicated(keep=False)]["Nome"].tolist()
    if duplicated:
        names = ", ".join(sorted(set(duplicated)))
        raise ValueError(f"Ha nomes duplicados na planilha: {names}")

    return df.sort_values("Antiguidade").reset_index(drop=True)


def read_uploaded_file(uploaded_file: Any) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    if name.endswith(".xlsx"):
        return pd.read_excel(uploaded_file)
    raise ValueError("Formato nao suportado. Importe um arquivo .xlsx ou .csv.")


def build_initial_state(df: pd.DataFrame, first_criterion: str, total_slots: int) -> dict[str, Any]:
    records = df.to_dict("records")
    by_name = {row["Nome"]: row for row in records}
    seniority_queue = [
        row["Nome"] for row in sorted(records, key=lambda item: (item["Antiguidade"], item["Nome"]))
    ]
    merit_queue = [
        row["Nome"] for row in sorted(records, key=lambda item: (item["Merecimento"], item["Nome"]))
    ]
    criteria = []
    current = first_criterion
    for _ in range(total_slots):
        criteria.append(current)
        current = "Merecimento" if current == "Antiguidade" else "Antiguidade"

    return {
        "df": df,
        "by_name": by_name,
        "seniority_queue": seniority_queue,
        "merit_queue": merit_queue,
        "criteria": criteria,
        "first_criterion": first_criterion,
        "total_slots": total_slots,
        "current_slot": 1,
        "consumed_slots": 0,
        "promotions": [],
        "history": [],
        "merit_round_count": 0,
        "merit_carryover": None,
        "current_round": None,
        "finished": False,
    }


def current_criterion() -> str | None:
    state = st.session_state.process
    if state["consumed_slots"] >= state["total_slots"]:
        return None
    return state["criteria"][state["consumed_slots"]]


def remove_from_queues(name: str) -> None:
    state = st.session_state.process
    state["seniority_queue"] = [item for item in state["seniority_queue"] if item != name]
    state["merit_queue"] = [item for item in state["merit_queue"] if item != name]
    if state["merit_carryover"] == name:
        state["merit_carryover"] = None


def register_promotion(
    name: str,
    criterion: str,
    consumed: bool,
    slot_number: int,
    observation: str,
) -> None:
    state = st.session_state.process
    row = state["by_name"][name]
    promotion = Promotion(
        ordem=len(state["promotions"]) + 1,
        nome=name,
        criterio=criterion,
        agregado="Sim" if row["Agregado"] else "Nao",
        consumiu_vaga="Sim" if consumed else "Nao",
        vaga_principal=slot_number,
        observacao=observation,
    )
    state["promotions"].append(asdict(promotion))
    remove_from_queues(name)


def pop_next_merit_name() -> str | None:
    state = st.session_state.process
    while state["merit_queue"]:
        name = state["merit_queue"].pop(0)
        if name not in state["by_name"]:
            continue
        already_promoted = any(item["nome"] == name for item in state["promotions"])
        if not already_promoted:
            return name
    return None


def process_seniority_slot() -> None:
    state = st.session_state.process
    slot_number = state["current_slot"]
    events = []

    while state["seniority_queue"]:
        name = state["seniority_queue"][0]
        row = state["by_name"][name]
        if row["Agregado"]:
            register_promotion(
                name,
                "Antiguidade",
                False,
                slot_number,
                "Promovido por antiguidade como agregado; nao consumiu vaga.",
            )
            events.append(f"{name} era agregado e foi promovido sem consumir vaga.")
            continue

        register_promotion(
            name,
            "Antiguidade",
            True,
            slot_number,
            "Promovido por antiguidade; consumiu vaga.",
        )
        state["consumed_slots"] += 1
        state["current_slot"] += 1
        state["history"].append(
            {
                "vaga_principal": slot_number,
                "criterio": "Antiguidade",
                "disputantes": name,
                "promovido": name,
                "desceu": "",
                "foi_para_fim": "",
                "eventos_agregado": " | ".join(events),
            }
        )
        break
    else:
        state["finished"] = True
        st.warning("Nao ha militares suficientes na fila de antiguidade para completar a vaga.")


def prepare_merit_round() -> None:
    state = st.session_state.process
    slot_number = state["current_slot"]
    state["merit_round_count"] += 1
    required = 2 if state["merit_round_count"] == 1 else 3
    candidates: list[str] = []
    events: list[str] = []

    if required == 3 and state["merit_carryover"]:
        carryover = state["merit_carryover"]
        if not any(item["nome"] == carryover for item in state["promotions"]):
            candidates.append(carryover)
        state["merit_carryover"] = None

    while len(candidates) < required:
        name = pop_next_merit_name()
        if name is None:
            state["current_round"] = {
                "type": "error",
                "message": "Nao ha militares suficientes na fila de merecimento para formar a disputa.",
            }
            return

        row = state["by_name"][name]
        if row["Agregado"]:
            register_promotion(
                name,
                "Merecimento",
                False,
                slot_number,
                "Promovido por merecimento como agregado durante composicao da disputa; nao consumiu vaga.",
            )
            events.append(f"{name} era agregado e foi promovido sem entrar na disputa.")
            continue

        if name not in candidates:
            candidates.append(name)

    state["current_round"] = {
        "type": "merit",
        "slot_number": slot_number,
        "required": required,
        "candidates": candidates,
        "events": events,
        "winner": None,
    }


def finish_merit_round(winner: str, carryover: str | None = None) -> None:
    state = st.session_state.process
    round_data = state["current_round"]
    slot_number = round_data["slot_number"]
    candidates = list(round_data["candidates"])
    remaining = [name for name in candidates if name != winner]

    register_promotion(
        winner,
        "Merecimento",
        True,
        slot_number,
        "Promovido por merecimento; consumiu vaga.",
    )

    sent_to_end = ""
    chosen_carryover = ""
    if len(candidates) == 2:
        chosen_carryover = remaining[0] if remaining else ""
        state["merit_carryover"] = chosen_carryover or None
    else:
        chosen_carryover = carryover or ""
        sent_to_end = next((name for name in remaining if name != chosen_carryover), "")
        state["merit_carryover"] = chosen_carryover or None
        if sent_to_end:
            state["merit_queue"] = [name for name in state["merit_queue"] if name != sent_to_end]
            state["merit_queue"].append(sent_to_end)

    state["history"].append(
        {
            "vaga_principal": slot_number,
            "criterio": "Merecimento",
            "disputantes": ", ".join(candidates),
            "promovido": winner,
            "desceu": chosen_carryover,
            "foi_para_fim": sent_to_end,
            "eventos_agregado": " | ".join(round_data["events"]),
        }
    )
    state["consumed_slots"] += 1
    state["current_slot"] += 1
    state["current_round"] = None


def queue_dataframe(queue: list[str], order_column: str) -> pd.DataFrame:
    state = st.session_state.process
    rows = []
    for index, name in enumerate(queue, start=1):
        row = state["by_name"][name]
        rows.append(
            {
                "Fila": index,
                "Nome": name,
                order_column: int(row[order_column]),
                "Agregado": "Sim" if row["Agregado"] else "Nao",
            }
        )
    return pd.DataFrame(rows)


def dataframe_download(df: pd.DataFrame, label: str, filename: str) -> None:
    csv = df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(label, csv, filename, "text/csv", use_container_width=True)


def promoted_display_dataframe(promotions: list[dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(promotions)
    if df.empty:
        return pd.DataFrame(
            columns=[
                "Ordem",
                "Nome",
                "Promovido por",
                "Agregado",
                "Consumiu vaga principal",
                "Vaga principal",
                "Observacao",
            ]
        )

    return df.rename(
        columns={
            "ordem": "Ordem",
            "nome": "Nome",
            "criterio": "Promovido por",
            "agregado": "Agregado",
            "consumiu_vaga": "Consumiu vaga principal",
            "vaga_principal": "Vaga principal",
            "observacao": "Observacao",
        }
    )[
        [
            "Ordem",
            "Nome",
            "Promovido por",
            "Agregado",
            "Consumiu vaga principal",
            "Vaga principal",
            "Observacao",
        ]
    ]


def style_excel_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def export_excel_bytes() -> bytes:
    state = st.session_state.process
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        promoted_display_dataframe(state["promotions"]).to_excel(writer, index=False, sheet_name="Promovidos")
        pd.DataFrame(state["history"]).to_excel(writer, index=False, sheet_name="Historico")
        queue_dataframe(state["seniority_queue"], "Antiguidade").to_excel(
            writer, index=False, sheet_name="Fila Antiguidade"
        )
        queue_dataframe(state["merit_queue"], "Merecimento").to_excel(
            writer, index=False, sheet_name="Fila Merecimento"
        )
        for sheet_name in writer.sheets:
            style_excel_sheet(writer, sheet_name)
    return output.getvalue()


def render_sidebar() -> None:
    st.sidebar.header("Configuracao")
    uploaded = st.sidebar.file_uploader("Importar planilha", type=["xlsx", "csv"])

    if uploaded:
        try:
            df = normalize_dataframe(read_uploaded_file(uploaded))
            st.session_state.loaded_df = df
            st.sidebar.success(f"{len(df)} militares carregados.")
        except Exception as exc:
            st.sidebar.error(str(exc))

    if "loaded_df" not in st.session_state:
        st.sidebar.info("Importe uma planilha .xlsx ou .csv para iniciar.")
        return

    first = st.sidebar.radio(
        "Primeira vaga",
        ["Antiguidade", "Merecimento"],
        horizontal=False,
    )
    total = st.sidebar.number_input("Vagas principais", min_value=1, step=1, value=1)

    if st.sidebar.button("Iniciar / reiniciar processamento", type="primary", use_container_width=True):
        st.session_state.process = build_initial_state(st.session_state.loaded_df, first, int(total))
        st.rerun()


def render_loaded_data() -> None:
    if "loaded_df" not in st.session_state:
        st.info("Carregue uma planilha para visualizar e processar as promocoes.")
        return
    shown = st.session_state.loaded_df.copy()
    shown["Agregado"] = shown["Agregado"].map({True: "Sim", False: "Nao"})
    st.subheader("Planilha carregada")
    st.dataframe(shown, hide_index=True, use_container_width=True)


def render_round_panel() -> None:
    if "process" not in st.session_state:
        return

    state = st.session_state.process
    if state["consumed_slots"] >= state["total_slots"]:
        state["finished"] = True

    st.subheader("Rodada atual")
    progress_text = f"Vagas principais consumidas: {state['consumed_slots']} de {state['total_slots']}"
    st.progress(state["consumed_slots"] / state["total_slots"], text=progress_text)

    if state["finished"]:
        st.success("Processamento concluido.")
        return

    criterion = current_criterion()
    st.caption(f"Vaga principal {state['current_slot']} - criterio: {criterion}")

    if criterion == "Antiguidade":
        st.write("A antiguidade e automatica: o sistema promove o primeiro elegivel da fila.")
        if st.button("Processar vaga por antiguidade", type="primary"):
            process_seniority_slot()
            st.rerun()
        return

    if state["current_round"] is None:
        st.write("No merecimento, o sistema monta a disputa e o usuario decide por cliques.")
        if st.button("Montar disputa por merecimento", type="primary"):
            prepare_merit_round()
            st.rerun()
        return

    round_data = state["current_round"]
    if round_data["type"] == "error":
        st.error(round_data["message"])
        return

    if round_data["events"]:
        st.info("Agregados promovidos durante a composicao: " + " | ".join(round_data["events"]))

    candidates = round_data["candidates"]
    st.write(f"Disputa formada com {len(candidates)} nao agregados.")
    cols = st.columns(len(candidates))
    for col, name in zip(cols, candidates):
        row = state["by_name"][name]
        with col:
            st.metric(name, f"Merecimento {int(row['Merecimento'])}")
            if st.button("Promover", key=f"promote_{round_data['slot_number']}_{name}", use_container_width=True):
                round_data["winner"] = name
                if len(candidates) == 2:
                    finish_merit_round(name)
                    st.rerun()

    if len(candidates) == 3 and round_data["winner"]:
        winner = round_data["winner"]
        remaining = [name for name in candidates if name != winner]
        st.divider()
        st.write(f"Promovido escolhido: **{winner}**. Agora selecione quem desce para a proxima disputa.")
        cols = st.columns(2)
        for col, name in zip(cols, remaining):
            with col:
                row = state["by_name"][name]
                st.metric(name, f"Merecimento {int(row['Merecimento'])}")
                if st.button(
                    f"Desce: {name}",
                    key=f"carry_{round_data['slot_number']}_{name}",
                    use_container_width=True,
                ):
                    finish_merit_round(winner, name)
                    st.rerun()


def render_process_tables() -> None:
    if "process" not in st.session_state:
        return
    state = st.session_state.process

    left, right = st.columns(2)
    with left:
        st.subheader("Fila de antiguidade")
        st.dataframe(
            queue_dataframe(state["seniority_queue"], "Antiguidade"),
            hide_index=True,
            use_container_width=True,
            height=320,
        )
    with right:
        st.subheader("Fila de merecimento")
        st.dataframe(
            queue_dataframe(state["merit_queue"], "Merecimento"),
            hide_index=True,
            use_container_width=True,
            height=320,
        )

    st.subheader("Lista unica de promovidos")
    promotions_df = promoted_display_dataframe(state["promotions"])
    st.dataframe(promotions_df, hide_index=True, use_container_width=True)

    st.subheader("Historico das rodadas")
    history_df = pd.DataFrame(state["history"])
    st.dataframe(history_df, hide_index=True, use_container_width=True)

    if not promotions_df.empty:
        col1, col2, col3 = st.columns(3)
        with col1:
            dataframe_download(promotions_df, "Exportar promovidos CSV", "promovidos.csv")
        with col2:
            dataframe_download(history_df, "Exportar historico CSV", "historico_rodadas.csv")
        with col3:
            st.download_button(
                "Exportar tudo XLSX",
                export_excel_bytes(),
                "promocoes_militares.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


def render_sample_download() -> None:
    sample = pd.DataFrame(
        [
            {"Nome": "Militar A", "Antiguidade": 1, "Merecimento": 4, "Agregado": "Nao"},
            {"Nome": "Militar B", "Antiguidade": 2, "Merecimento": 1, "Agregado": "Sim"},
            {"Nome": "Militar C", "Antiguidade": 3, "Merecimento": 2, "Agregado": "Nao"},
            {"Nome": "Militar D", "Antiguidade": 4, "Merecimento": 3, "Agregado": "Nao"},
            {"Nome": "Militar E", "Antiguidade": 5, "Merecimento": 5, "Agregado": "Nao"},
            {"Nome": "Militar F", "Antiguidade": 6, "Merecimento": 6, "Agregado": "Nao"},
            {"Nome": "Militar G", "Antiguidade": 7, "Merecimento": 7, "Agregado": "Nao"},
            {"Nome": "Militar H", "Antiguidade": 8, "Merecimento": 8, "Agregado": "Sim"},
            {"Nome": "Militar I", "Antiguidade": 9, "Merecimento": 9, "Agregado": "Nao"},
            {"Nome": "Militar J", "Antiguidade": 10, "Merecimento": 10, "Agregado": "Nao"},
        ]
    )
    st.sidebar.download_button(
        "Baixar CSV de exemplo",
        sample.to_csv(index=False).encode("utf-8-sig"),
        "modelo_militares.csv",
        "text/csv",
        use_container_width=True,
    )


def main() -> None:
    st.set_page_config(page_title="Promocoes militares", layout="wide")
    st.title("Gerenciador de promocoes por antiguidade e merecimento")
    st.caption("Importe a planilha, configure as vagas e processe cada rodada por cliques.")

    render_sidebar()
    render_sample_download()
    render_loaded_data()
    render_round_panel()
    render_process_tables()


if __name__ == "__main__":
    main()
